import shutil
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def exportar_datos(excel_madre_path, salida_path, df_s400=None, df_fsat=None,
                   hoja_s400='s400', hoja_fsat='fsat', confirmar_sobrescritura=False):
    try:
        # Verificar que el archivo base existe
        if not os.path.exists(excel_madre_path):
            print(f"ERROR: El archivo base '{excel_madre_path}' no existe.")
            return

        # Si el archivo de salida ya existe, preguntar si sobrescribir
        if os.path.exists(salida_path):
            if confirmar_sobrescritura:
                respuesta = input(f"El archivo '{salida_path}' ya existe. ¿Deseas sobrescribirlo? (s/n): ")
                if respuesta.lower() != 's':
                    print("Exportación cancelada por el usuario.")
                    return
            else:
                print(f"Advertencia: El archivo '{salida_path}' será sobrescrito automáticamente.")

        # Copiar el archivo base
        shutil.copyfile(excel_madre_path, salida_path)
        print(f"Archivo base copiado a: {salida_path}")

        # Cargar el libro copiado
        wb = load_workbook(salida_path)

        # Insertar datos de s400
        if df_s400 is not None:
            if hoja_s400 in wb.sheetnames:
                del wb[hoja_s400]
            ws_s400 = wb.create_sheet(hoja_s400)
            for r in dataframe_to_rows(df_s400, index=False, header=True):
                ws_s400.append(r)
            print(f"Se exportaron {len(df_s400)} filas a la hoja '{hoja_s400}'.")

        # Insertar datos de fsat
        if df_fsat is not None:
            if hoja_fsat in wb.sheetnames:
                del wb[hoja_fsat]
            ws_fsat = wb.create_sheet(hoja_fsat)
            for r in dataframe_to_rows(df_fsat, index=False, header=True):
                ws_fsat.append(r)
            print(f"Se exportaron {len(df_fsat)} filas a la hoja '{hoja_fsat}'.")

        # Guardar cambios
        wb.save(salida_path)
        print(f"\n¡Datos exportados correctamente a:\n{salida_path}")

    except Exception as e:
        print(f"Error al exportar datos: {e}")